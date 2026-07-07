"""
Plugin: Reporte Excel Automático — LBAMonitor v4.2
==================================================
Genera un reporte Excel al cerrar sesión de servicio.
Guarda en C:/ProgramData/LBAMonitor/exports/reports/

Requiere: openpyxl (incluido en pyproject.toml)

Columnas del reporte:
  - Fecha, Dispositivo, Archivos copiados, Cobro (CUP)
"""
import os
from datetime import datetime
from pathlib import Path

def on_session_ended(session_id: int) -> None:
    """Genera reporte Excel al cerrar sesión del día."""
    _generate_daily_report()


def _generate_daily_report() -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return  # openpyxl no disponible

    try:
        # Importar aquí para evitar circular imports al cargar el plugin
        # El plugin se carga DESPUÉS de que el ORM está inicializado
        exports_dir = Path(
            os.environ.get("LBAMONITOR_EXPORTS", "C:/ProgramData/LBAMonitor/exports/reports")
        )
        exports_dir.mkdir(parents=True, exist_ok=True)

        today = datetime.now().strftime("%Y-%m-%d")
        filename = exports_dir / f"reporte_{today}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {today}"

        # Encabezado
        headers = ["Fecha/Hora", "Dispositivo", "Etiqueta", "Archivos Copiados",
                   "Bytes Copiados", "Cobro (CUP)", "Cliente"]
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="F9FAFB", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Intentar obtener datos reales de la BD
        try:
            _fill_from_db(ws, today)
        except Exception:
            # Sin datos BD disponibles, dejar el header vacío
            pass

        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(filename)
        print(f"[excel_report] Reporte guardado: {filename}")
    except Exception as e:
        print(f"[excel_report] Error generando reporte: {e}")


def _fill_from_db(ws, today: str) -> None:
    """Intenta rellenar el reporte con datos de la BD (best-effort)."""
    import sqlite3, os
    db_path = os.environ.get(
        "LBAMONITOR_DB", "C:/ProgramData/LBAMonitor/data/lbamonitor.db"
    )
    if not os.path.exists(db_path):
        return
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.execute("""
            SELECT
                d.insertion_date_time,
                d.name,
                d.volume_label,
                COUNT(c.id) as files,
                COALESCE(SUM(c.size_bytes), 0) as bytes,
                COALESCE(d.payment, 0) as payment,
                COALESCE(cl.name, '') as client
            FROM inserted_drives d
            LEFT JOIN copies c ON c.inserted_drive_id = d.id
            LEFT JOIN usb_devices ud ON ud.id = d.usb_device_id
            LEFT JOIN clients cl ON cl.device_id = ud.id
            WHERE date(d.insertion_date_time) = ?
            GROUP BY d.id
            ORDER BY d.insertion_date_time
        """, (today,))
        for row_idx, row in enumerate(cur.fetchall(), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
    finally:
        conn.close()
